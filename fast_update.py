"""
Fast bulk update using OSRM Table API.
Reads Excel, geocodes, computes full driving matrix in batch, saves all.
"""
import openpyxl
import json
import time
import urllib.request
import urllib.parse
import os
import sys

EXCEL_PATH = r"C:\Users\rasmu\Claude arbejde\Diverse til kodearbejde\Kørselsmatrix med alle fodboldklubber på fyn.xlsx"
CACHE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "Kørselstid mellem klubber program", "cache")
CACHE_PATH = os.path.join(CACHE_DIR, "geocode_cache.json")

KNOWN_COORDS = {
    "Aarslev BK": (55.308319, 10.490157),
    "DBU Fyn": (55.3954, 10.3516),
    "F.C. Lange Bolde": (55.4120, 10.3640),
    "FC Avrasya": (55.3990, 10.3970),
    "FC Hjallese": (55.3970, 10.3900),
    "FIUK, Odense": (55.4200, 10.4250),
    "Get2Sport": (55.4120, 10.4130),
    "Rise S & IF": (54.8850, 10.3730),
}

# === Step 1: Read Excel ===
print("=" * 60)
print("Step 1: Reading Excel")
print("=" * 60)
wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb.active

clubs = []
for row in range(2, ws.max_row + 1):
    name = ws.cell(row, 1).value
    if not name or not str(name).strip():
        continue
    addr = ws.cell(row, 2).value or ''
    postal = ws.cell(row, 3).value or ''
    city = ws.cell(row, 4).value or ''
    clubs.append({
        'name': str(name).strip(),
        'address': str(addr).strip(),
        'postal_code': str(int(postal) if isinstance(postal, (int, float)) else postal).strip(),
        'city': str(city).strip()
    })
print(f"  {len(clubs)} clubs in Excel")

# === Step 2: Geocode ===
print("\n" + "=" * 60)
print("Step 2: Geocoding")
print("=" * 60)

geocode_cache = {}
if os.path.exists(CACHE_PATH):
    with open(CACHE_PATH, "r", encoding="utf-8") as f:
        geocode_cache = json.load(f)

def geocode_nominatim(address):
    params = urllib.parse.urlencode({
        "q": address, "format": "json", "limit": 1, "countrycodes": "dk"
    })
    url = f"https://nominatim.openstreetmap.org/search?{params}"
    req = urllib.request.Request(url, headers={"User-Agent": "KoerselstidFodbold/1.0"})
    with urllib.request.urlopen(req, timeout=15) as resp:
        data = json.loads(resp.read().decode())
    if data:
        return float(data[0]['lat']), float(data[0]['lon']), data[0].get('display_name', '')
    return None, None, None

coords = {}  # name -> (lat, lon)
need_geocoding = []

for c in clubs:
    name = c['name']
    if name in KNOWN_COORDS:
        coords[name] = KNOWN_COORDS[name]
    elif name in geocode_cache:
        coords[name] = (float(geocode_cache[name]['lat']), float(geocode_cache[name]['lon']))
    else:
        need_geocoding.append(c)

print(f"  {len(coords)} from cache/known, {len(need_geocoding)} need geocoding")

for c in need_geocoding:
    name = c['name']
    addr = f"{c['address']}, {c['postal_code']} {c['city']}, Danmark"
    lat, lon, display = geocode_nominatim(addr)
    if not lat:
        addr2 = f"{c['address']}, {c['city']}, Danmark"
        lat, lon, display = geocode_nominatim(addr2)
    if lat:
        coords[name] = (lat, lon)
        geocode_cache[name] = {"lat": str(lat), "lon": str(lon), "display_name": display or ""}
        print(f"  {name}: {lat:.4f}, {lon:.4f}")
    else:
        print(f"  WARNING: Could not geocode {name}")
    time.sleep(1.1)

# Save geocode cache
os.makedirs(os.path.dirname(CACHE_PATH), exist_ok=True)
with open(CACHE_PATH, "w", encoding="utf-8") as f:
    json.dump(geocode_cache, f, ensure_ascii=False, indent=2)

clubs_with_coords = [c for c in clubs if c['name'] in coords]
names = [c['name'] for c in clubs_with_coords]
lats = [coords[c['name']][0] for c in clubs_with_coords]
lons = [coords[c['name']][1] for c in clubs_with_coords]
print(f"\n  {len(names)} clubs with coordinates")

# === Step 3: OSRM Table API ===
print("\n" + "=" * 60)
print("Step 3: Calculating routes (OSRM Table API)")
print("=" * 60)

# Build coordinate string for all clubs
coord_str = ";".join(f"{lon},{lat}" for lon, lat in zip(lons, lats))
base_url = f"http://router.project-osrm.org/table/v1/driving/{coord_str}"

n = len(names)
BATCH = 40  # sources per batch (keep URL manageable)
matrix = {}
total_entries = 0

# Process in row batches (source batches)
num_batches = (n + BATCH - 1) // BATCH
for batch_idx in range(num_batches):
    start = batch_idx * BATCH
    end = min(start + BATCH, n)
    sources = ";".join(str(i) for i in range(start, end))
    url = f"{base_url}?sources={sources}&annotations=duration,distance"

    print(f"\n  Batch {batch_idx + 1}/{num_batches}: sources {start}-{end-1} ({end-start} clubs) vs all {n} destinations...")
    print(f"  URL length: {len(url)} chars")

    success = False
    for attempt in range(3):
        try:
            req = urllib.request.Request(url, headers={"User-Agent": "KoerselstidFodbold/1.0"})
            with urllib.request.urlopen(req, timeout=120) as resp:
                data = json.loads(resp.read().decode())

            if data["code"] != "Ok":
                print(f"    Error: {data.get('message', 'Unknown')}")
                break

            durations = data["durations"]
            distances = data["distances"]
            batch_entries = 0

            for si, src_idx in enumerate(range(start, end)):
                src_name = names[src_idx]
                for dst_idx in range(n):
                    dst_name = names[dst_idx]
                    dur_sec = durations[si][dst_idx]
                    dist_m = distances[si][dst_idx]

                    if dur_sec is None or dist_m is None:
                        continue

                    dur_min = round(dur_sec / 60)
                    dist_km = round(dist_m / 1000, 1)
                    mins = int(dur_sec // 60)
                    if mins >= 60:
                        hours = mins // 60
                        remaining = mins % 60
                        dur_text = f"{hours} t {remaining} min"
                    else:
                        dur_text = f"{mins} min"

                    matrix[f"{src_name}|{dst_name}"] = {
                        "duration_min": dur_min,
                        "duration_sec": round(dur_sec),
                        "distance_km": dist_km,
                        "duration_text": dur_text
                    }
                    batch_entries += 1

            total_entries += batch_entries
            print(f"    OK: {batch_entries} entries (total: {total_entries})")
            success = True
            break

        except Exception as e:
            if attempt < 2:
                print(f"    Retry {attempt+1}: {e}")
                time.sleep(10 * (attempt + 1))
            else:
                print(f"    FAILED after 3 attempts: {e}")

    if not success:
        print(f"    WARNING: Batch {batch_idx + 1} failed!")

    time.sleep(2)  # Rate limit between batches

print(f"\n  Total matrix entries: {total_entries}")

# === Step 4: Save ===
print("\n" + "=" * 60)
print("Step 4: Saving data")
print("=" * 60)

with open("data/clubs.json", "w", encoding="utf-8") as f:
    json.dump(clubs, f, ensure_ascii=False, indent=2)
print(f"  Saved data/clubs.json ({len(clubs)} clubs)")

with open("data/matrix.json", "w", encoding="utf-8") as f:
    json.dump(matrix, f, ensure_ascii=False)
print(f"  Saved data/matrix.json ({len(matrix)} entries)")

# Verify
missing_clubs = [c['name'] for c in clubs if c['name'] not in coords]
if missing_clubs:
    print(f"\n  WARNING: {len(missing_clubs)} clubs without coordinates:")
    for name in missing_clubs:
        print(f"    - {name}")

expected = len(names) * len(names)
print(f"\n  Expected entries: {expected}, Got: {len(matrix)}")
if len(matrix) < expected * 0.95:
    print("  WARNING: More than 5% routes missing!")

print("\nDone! Now run generate_exports.py")
