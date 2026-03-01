"""
Full update: read new Excel, geocode new/changed clubs, recalculate all
affected routes via OSRM, update matrix.json and clubs.json.
"""
import openpyxl
import json
import time
import urllib.request
import urllib.parse
import os

EXCEL_PATH = r"C:\Users\rasmu\Claude arbejde\Diverse til kodearbejde\Kørselsmatrix med alle fodboldklubber på fyn.xlsx"

# === Step 1: Read new Excel ===
print("=" * 60)
print("Step 1: Reading new Excel file")
print("=" * 60)
wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb.active

new_clubs = []
for row in range(2, ws.max_row + 1):
    name = ws.cell(row, 1).value
    if not name or not str(name).strip():
        continue
    addr = ws.cell(row, 2).value or ''
    postal = ws.cell(row, 3).value or ''
    city = ws.cell(row, 4).value or ''
    new_clubs.append({
        'name': str(name).strip(),
        'address': str(addr).strip(),
        'postal_code': str(int(postal) if isinstance(postal, (int, float)) else postal).strip(),
        'city': str(city).strip()
    })

print(f"  {len(new_clubs)} clubs in new Excel")

# === Step 2: Load existing data ===
with open("data/clubs.json", "r", encoding="utf-8") as f:
    old_clubs = json.load(f)

with open("data/matrix.json", "r", encoding="utf-8") as f:
    matrix = json.load(f)

# Load geocode cache
cache_path = os.path.join(os.path.dirname(__file__), "..", "Kørselstid mellem klubber program", "cache", "geocode_cache.json")
geocode_cache = {}
if os.path.exists(cache_path):
    with open(cache_path, "r", encoding="utf-8") as f:
        geocode_cache = json.load(f)

old_map = {c['name']: c for c in old_clubs}
new_map = {c['name']: c for c in new_clubs}

# Find clubs that need geocoding (new or changed address)
needs_geocoding = []
for c in new_clubs:
    if c['name'] not in old_map:
        needs_geocoding.append(c)
    elif (c['address'] != old_map[c['name']]['address'] or
          c['postal_code'] != old_map[c['name']]['postal_code'] or
          c['city'] != old_map[c['name']]['city']):
        needs_geocoding.append(c)

print(f"  {len(needs_geocoding)} clubs need geocoding")

# === Step 3: Geocode ===
print("\n" + "=" * 60)
print("Step 2: Geocoding new/changed clubs")
print("=" * 60)

# Hardcoded coords for known problem cases
KNOWN_COORDS = {
    "Aarslev BK": (55.308319, 10.490157),  # Fyn, not Aarhus
    "DBU Fyn": (55.3954, 10.3516),  # Stadionvej 50, 5200 Odense V (same as BK2020)
    "F.C. Lange Bolde": (55.4120, 10.3640),  # Rugårdsvej 242, 5210 Odense NV
    "FC Avrasya": (55.3990, 10.3970),  # Risingvej 25, 5000 Odense C
    "FC Hjallese": (55.3970, 10.3900),  # Schacksgade 14, 5000 Odense C
    "FIUK, Odense": (55.4200, 10.4250),  # Risingevej 122, 5240 Odense NØ
    "Get2Sport": (55.4120, 10.4130),  # Vollsmose Alle 20, 5240 Odense NØ
    "Rise S & IF": (54.8850, 10.3730),  # St. Rise Skolevej, 5970 Ærøskøbing
}

coords = {}

# Load existing coords from cache
for c in new_clubs:
    name = c['name']
    if name in KNOWN_COORDS:
        coords[name] = KNOWN_COORDS[name]
    elif name in geocode_cache and name not in [n['name'] for n in needs_geocoding]:
        coords[name] = (float(geocode_cache[name]['lat']), float(geocode_cache[name]['lon']))

def geocode(address):
    params = urllib.parse.urlencode({
        "q": address, "format": "json", "limit": 1, "countrycodes": "dk"
    })
    url = f"https://nominatim.openstreetmap.org/search?{params}"
    req = urllib.request.Request(url, headers={"User-Agent": "KoerselstidFodbold/1.0"})
    with urllib.request.urlopen(req) as resp:
        data = json.loads(resp.read().decode())
    if data:
        return float(data[0]['lat']), float(data[0]['lon']), data[0].get('display_name', '')
    return None, None, None

for c in needs_geocoding:
    name = c['name']
    if name in KNOWN_COORDS:
        coords[name] = KNOWN_COORDS[name]
        print(f"  {name}: using hardcoded coords")
        continue

    addr = f"{c['address']}, {c['postal_code']} {c['city']}, Danmark"
    lat, lon, display = geocode(addr)
    if lat:
        coords[name] = (lat, lon)
        # Update cache
        geocode_cache[name] = {"lat": str(lat), "lon": str(lon), "display_name": display}
        print(f"  {name}: {lat:.4f}, {lon:.4f} ({display[:60]})")
    else:
        # Try with just postal code and city
        addr2 = f"{c['address']}, {c['city']}, Danmark"
        lat, lon, display = geocode(addr2)
        if lat:
            coords[name] = (lat, lon)
            geocode_cache[name] = {"lat": str(lat), "lon": str(lon), "display_name": display}
            print(f"  {name}: {lat:.4f}, {lon:.4f} (fallback: {display[:60]})")
        else:
            print(f"  WARNING: Could not geocode {name} ({addr})")
    time.sleep(1.1)  # Nominatim rate limit

# Save updated geocode cache
with open(cache_path, "w", encoding="utf-8") as f:
    json.dump(geocode_cache, f, ensure_ascii=False, indent=2)

print(f"\n  Total coords: {len(coords)}/{len(new_clubs)}")

# === Step 4: Calculate routes ===
print("\n" + "=" * 60)
print("Step 3: Calculating routes via OSRM")
print("=" * 60)

def get_route(lat1, lon1, lat2, lon2, retries=3):
    url = f"http://router.project-osrm.org/route/v1/driving/{lon1},{lat1};{lon2},{lat2}?overview=false"
    for attempt in range(retries):
        try:
            req = urllib.request.Request(url, headers={"User-Agent": "KoerselstidFodbold/1.0"})
            with urllib.request.urlopen(req, timeout=30) as resp:
                data = json.loads(resp.read().decode())
            if data["code"] == "Ok" and data["routes"]:
                route = data["routes"][0]
                duration_sec = route["duration"]
                distance_m = route["distance"]
                duration_min = round(duration_sec / 60)
                distance_km = round(distance_m / 1000, 1)
                mins = int(duration_sec // 60)
                if mins >= 60:
                    hours = mins // 60
                    remaining = mins % 60
                    duration_text = f"{hours} t {remaining} min"
                else:
                    duration_text = f"{mins} min"
                return {
                    "duration_min": duration_min,
                    "duration_sec": round(duration_sec),
                    "distance_km": distance_km,
                    "duration_text": duration_text
                }
        except Exception as e:
            if attempt < retries - 1:
                time.sleep(5 * (attempt + 1))
            else:
                print(f"    Route error after {retries} attempts: {e}")
    return None

# Determine which routes need calculation
# Routes involving new/changed clubs need recalculation
clubs_needing_routes = set(c['name'] for c in needs_geocoding)

# Also remove routes for clubs no longer in the list
removed_clubs = set(c['name'] for c in old_clubs if c['name'] not in new_map)
if removed_clubs:
    print(f"  Removing routes for {len(removed_clubs)} removed clubs")
    keys_to_remove = [k for k in matrix if any(rc in k for rc in removed_clubs)]
    for k in keys_to_remove:
        del matrix[k]
    print(f"  Removed {len(keys_to_remove)} routes")

# Calculate all routes for clubs needing routes
all_names = [c['name'] for c in new_clubs if c['name'] in coords]
total_routes = 0
new_routes = 0
errors = 0

for club_name in clubs_needing_routes:
    if club_name not in coords:
        print(f"  Skipping {club_name} (no coordinates)")
        continue

    lat1, lon1 = coords[club_name]
    print(f"\n  Calculating routes for {club_name}...")

    for other_name in all_names:
        if other_name == club_name:
            # Self route
            matrix[f"{club_name}|{club_name}"] = {
                "duration_min": 0, "duration_sec": 0,
                "distance_km": 0.0, "duration_text": "0 min"
            }
            continue

        if other_name not in coords:
            continue

        lat2, lon2 = coords[other_name]

        # Club -> Other
        key1 = f"{club_name}|{other_name}"
        route1 = get_route(lat1, lon1, lat2, lon2)
        if route1:
            matrix[key1] = route1
            new_routes += 1
        else:
            errors += 1
        total_routes += 1
        time.sleep(0.6)

        # Other -> Club
        key2 = f"{other_name}|{club_name}"
        route2 = get_route(lat2, lon2, lat1, lon1)
        if route2:
            matrix[key2] = route2
            new_routes += 1
        else:
            errors += 1
        total_routes += 1
        time.sleep(0.6)

        if total_routes % 50 == 0:
            print(f"    ... {total_routes} routes calculated ({new_routes} ok, {errors} errors)")

    print(f"    Done ({new_routes} routes so far)")

print(f"\n  Total: {new_routes} new routes, {errors} errors")
print(f"  Matrix size: {len(matrix)} entries")

# === Step 5: Save ===
print("\n" + "=" * 60)
print("Step 4: Saving updated data")
print("=" * 60)

# Save clubs.json
with open("data/clubs.json", "w", encoding="utf-8") as f:
    json.dump(new_clubs, f, ensure_ascii=False, indent=2)
print(f"  Saved data/clubs.json ({len(new_clubs)} clubs)")

# Save matrix.json
with open("data/matrix.json", "w", encoding="utf-8") as f:
    json.dump(matrix, f, ensure_ascii=False)
print(f"  Saved data/matrix.json ({len(matrix)} entries)")

print("\nDone! Now run generate_exports.py to create Excel/CSV files.")
