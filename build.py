"""
Build script for koerselstid-fodbold static site.
Regenerates data/clubs.json from the source Excel file.

Usage:
    python build.py --excel path/to/klubber.xlsx

Optional:
    python build.py --excel path/to/klubber.xlsx --matrix path/to/driving_matrix.json
"""
import json
import argparse
from pathlib import Path

def read_excel(path):
    """Read club data from Excel file."""
    import openpyxl
    wb = openpyxl.load_workbook(str(path))
    ws = wb["Ark1"]
    clubs = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=4, values_only=True):
        if row[0]:
            clubs.append({
                "name": str(row[0]).strip(),
                "address": str(row[1]).strip() if row[1] else "",
                "postal_code": str(row[2]).strip() if row[2] else "",
                "city": str(row[3]).strip() if row[3] else ""
            })
    return clubs

def main():
    parser = argparse.ArgumentParser(description="Build static data files for koerselstid-fodbold")
    parser.add_argument("--excel", required=True, help="Path to klubber.xlsx")
    parser.add_argument("--matrix", help="Path to driving_matrix.json (optional, copies to data/matrix.json)")
    args = parser.parse_args()

    base_dir = Path(__file__).parent
    data_dir = base_dir / "data"
    data_dir.mkdir(exist_ok=True)

    # Generate clubs.json
    clubs = read_excel(args.excel)
    clubs_path = data_dir / "clubs.json"
    with open(clubs_path, "w", encoding="utf-8") as f:
        json.dump(clubs, f, ensure_ascii=False, indent=2)
    print(f"Generated {clubs_path} with {len(clubs)} clubs")

    # Optionally copy matrix
    if args.matrix:
        import shutil
        matrix_dest = data_dir / "matrix.json"
        shutil.copy2(args.matrix, matrix_dest)
        with open(matrix_dest, "r", encoding="utf-8") as f:
            matrix = json.load(f)
        print(f"Copied matrix to {matrix_dest} ({len(matrix)} entries)")

    print("Done!")

if __name__ == "__main__":
    main()
