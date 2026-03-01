"""Generate Excel and CSV export files from matrix.json and clubs.json."""
import json
import csv
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Load data
with open("data/clubs.json", "r", encoding="utf-8") as f:
    clubs_data = json.load(f)

with open("data/matrix.json", "r", encoding="utf-8") as f:
    driving_matrix = json.load(f)

club_names = sorted([c["name"] for c in clubs_data])
print(f"Loaded {len(club_names)} clubs and {len(driving_matrix)} routes")

# === Generate CSV ===
output = io.StringIO()
writer = csv.writer(output, delimiter=";")
writer.writerow(["Klub"] + club_names)
for cn1 in club_names:
    row = [cn1]
    for cn2 in club_names:
        key = f"{cn1}|{cn2}"
        if key in driving_matrix:
            row.append(driving_matrix[key]["duration_min"])
        else:
            row.append("")
    writer.writerow(row)

with open("exports/koerselstider_matrix.csv", "w", encoding="utf-8-sig", newline="") as f:
    f.write(output.getvalue())
print("Generated exports/koerselstider_matrix.csv")

# === Generate Excel ===
wb = openpyxl.Workbook()

header_font = Font(bold=True, size=10, color="FFFFFF")
header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
cell_alignment = Alignment(horizontal="center", vertical="center")
thin_border = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0')
)

green_light = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
yellow_light = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
orange_light = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")
red_light = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
gray_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

# Sheet 1: Matrix (min)
ws = wb.active
ws.title = "Kørselstid Matrix (min)"

ws.cell(row=1, column=1, value="Klub").font = header_font
ws.cell(row=1, column=1).fill = header_fill
ws.cell(row=1, column=1).alignment = cell_alignment

for j, name in enumerate(club_names):
    cell = ws.cell(row=1, column=j+2, value=name)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", text_rotation=90)

for i, cn1 in enumerate(club_names):
    name_cell = ws.cell(row=i+2, column=1, value=cn1)
    name_cell.font = Font(bold=True, size=9)
    name_cell.fill = green_light

    for j, cn2 in enumerate(club_names):
        key = f"{cn1}|{cn2}"
        cell = ws.cell(row=i+2, column=j+2)
        cell.alignment = cell_alignment
        cell.border = thin_border

        if cn1 == cn2:
            cell.value = 0
            cell.fill = gray_fill
        elif key in driving_matrix:
            mins = driving_matrix[key]["duration_min"]
            cell.value = mins
            if mins <= 15:
                cell.fill = green_light
            elif mins <= 30:
                cell.fill = yellow_light
            elif mins <= 45:
                cell.fill = orange_light
            else:
                cell.fill = red_light

ws.column_dimensions['A'].width = 25

# Sheet 2: Club details
ws2 = wb.create_sheet("Kluboversigt")
headers2 = ["Klubnavn", "Adresse", "Postnummer", "By"]
for j, h in enumerate(headers2):
    cell = ws2.cell(row=1, column=j+1, value=h)
    cell.font = header_font
    cell.fill = header_fill

for i, club in enumerate(clubs_data):
    ws2.cell(row=i+2, column=1, value=club["name"])
    ws2.cell(row=i+2, column=2, value=club["address"])
    ws2.cell(row=i+2, column=3, value=club["postal_code"])
    ws2.cell(row=i+2, column=4, value=club["city"])

for col_letter in ['A', 'B', 'C', 'D']:
    ws2.column_dimensions[col_letter].width = 30

# Sheet 3: Distance (km)
ws3 = wb.create_sheet("Afstand (km)")

ws3.cell(row=1, column=1, value="Klub").font = header_font
ws3.cell(row=1, column=1).fill = header_fill

for j, name in enumerate(club_names):
    cell = ws3.cell(row=1, column=j+2, value=name)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", text_rotation=90)

for i, cn1 in enumerate(club_names):
    ws3.cell(row=i+2, column=1, value=cn1).font = Font(bold=True, size=9)
    for j, cn2 in enumerate(club_names):
        key = f"{cn1}|{cn2}"
        cell = ws3.cell(row=i+2, column=j+2)
        cell.alignment = cell_alignment
        if key in driving_matrix:
            cell.value = driving_matrix[key]["distance_km"]

ws3.column_dimensions['A'].width = 25

wb.save("exports/koerselstider_matrix.xlsx")
print("Generated exports/koerselstider_matrix.xlsx")
print("Done!")
